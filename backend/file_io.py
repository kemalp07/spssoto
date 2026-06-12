"""Dosya okuma."""
import io
import math
import os
import tempfile
from collections import Counter
from typing import List, Dict, Any, Optional, Tuple
import numpy as np
import pandas as pd
from fastapi import HTTPException
from utils import sanitize
from missing_code_infer import infer_missing_codes_from_dataframe

def _cell_str(val) -> str:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return ""
    return str(val).strip()

def _parse_excel_rows(rows: List[list]) -> Tuple[List[dict], List[str], Dict[str, str]]:
    if not rows:
        raise HTTPException(status_code=400, detail="Boş dosya")

    header_row = [_cell_str(c) for c in rows[0]]
    labels: Dict[str, str] = {}
    data_start = 1

    if len(rows) > 1:
        second_row = [_cell_str(c) for c in rows[1]]
        is_label_row = (
            all(
                not v.replace(".", "").replace("-", "").isdigit()
                for v in second_row if v
            )
            and second_row != header_row
            and any(
                v and h and v.lower() != h.lower()
                for v, h in zip(second_row, header_row)
            )
        )
        if is_label_row:
            labels = {h: v for h, v in zip(header_row, second_row) if v and h}
            data_start = 2

    records = []
    for row in rows[data_start:]:
        if all(_cell_str(c) == "" for c in row):
            continue
        records.append({
            header_row[i]: row[i] if i < len(row) else ""
            for i in range(len(header_row))
            if header_row[i]
        })

    return records, header_row, labels


def read_uploaded_file(filename: str, file_bytes: bytes) -> dict:
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Dosya boş")

    labels: Dict[str, str] = {}
    value_labels: Dict[str, Any] = {}
    variable_measure: Dict[str, str] = {}
    missing_codes: Dict[str, List[str]] = {}
    global_missing_code: Optional[str] = None
    records: List[dict] = []
    columns: List[str] = []
    source = "excel"

    if filename.endswith(".sav"):
        try:
            import pyreadstat

            with tempfile.NamedTemporaryFile(suffix=".sav", delete=False) as f:
                f.write(file_bytes)
                tmp_path = f.name
            try:
                df, meta = pyreadstat.read_sav(tmp_path, user_missing=True)
            finally:
                os.unlink(tmp_path)

            if meta.column_names_to_labels:
                for col, lbl in meta.column_names_to_labels.items():
                    if lbl and str(lbl).strip() and str(lbl).strip() != col:
                        labels[str(col)] = str(lbl).strip()

            value_labels = meta.variable_value_labels or {}
            variable_measure = meta.variable_measure or {}

            missing_ranges = meta.missing_ranges or {}
            for col, ranges in missing_ranges.items():
                codes = []
                for r in ranges:
                    lo, hi = r.get("lo"), r.get("hi")
                    if lo is not None and hi is not None:
                        if lo == hi:
                            codes.append(
                                str(int(lo)) if float(lo).is_integer() else str(lo)
                            )
                        else:
                            codes.append(f"{lo}-{hi}")
                if codes:
                    missing_codes[col] = codes

            all_codes = [c for codes in missing_codes.values() for c in codes]
            if all_codes:
                global_missing_code = Counter(all_codes).most_common(1)[0][0]

            records = df.replace({np.nan: None}).to_dict(orient="records")
            columns = [str(c) for c in df.columns]
            source = "spss"

            inferred, inferred_global = infer_missing_codes_from_dataframe(df)
            for col, codes in inferred.items():
                existing = set(missing_codes.get(col, []))
                merged = sorted(existing | set(codes), key=lambda x: (len(x), x))
                if merged:
                    missing_codes[col] = merged
            if inferred_global and not global_missing_code:
                global_missing_code = inferred_global
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"SAV okunamadı: {e}")

    elif filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            records, columns, labels = _parse_excel_rows(rows)
            wb.close()
            if records:
                df = pd.DataFrame(records)
                inferred, inferred_global = infer_missing_codes_from_dataframe(df)
                missing_codes.update(inferred)
                if inferred_global:
                    global_missing_code = inferred_global
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Excel okunamadı: {e}")

    elif filename.endswith(".xls"):
        try:
            df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None)
            rows = df_raw.values.tolist()
            records, columns, labels = _parse_excel_rows(rows)
            if records:
                df = pd.DataFrame(records)
                inferred, inferred_global = infer_missing_codes_from_dataframe(df)
                missing_codes.update(inferred)
                if inferred_global:
                    global_missing_code = inferred_global
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Excel okunamadı: {e}")

    elif filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="CSV için frontend okuma kullanın")
    else:
        raise HTTPException(status_code=400, detail="Desteklenmeyen format")

    if not records:
        raise HTTPException(status_code=400, detail="Dosya boş görünüyor")

    return sanitize({
        "data": records,
        "labels": labels,
        "value_labels": {
            str(k): {str(vk): str(vv) for vk, vv in v.items()}
            for k, v in value_labels.items()
        },
        "variable_measure": variable_measure,
        "missing_codes": missing_codes,
        "global_missing_code": global_missing_code,
        "columns": columns,
        "row_count": len(records),
        "source": source,
        "labels_found": len(labels),
    })

